from openpyxl import load_workbook
import re
from collections import defaultdict

def trace_excel_formulas_full(file):
    wb = load_workbook(file, data_only=False)
    traces = {}
    hop_summary = defaultdict(int)

    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'f':
                    trace = trace_formula_chain_with_hops(wb, sheet_name, cell.coordinate)
                    traces[f"{sheet_name}!{cell.coordinate}"] = trace['lines']
                    hop_summary[len(trace['lines'])] += 1

    return traces, hop_summary


def trace_formula_chain_with_hops(wb, sheet_name, cell_ref, visited=None, depth=0, max_depth=500):
    if visited is None:
        visited = set()

    key = (sheet_name, cell_ref)
    if key in visited or depth > max_depth:
        return {'lines': [f"{'  ' * depth}{sheet_name}!{cell_ref} = [CIRCULAR or TOO DEEP]"], 'hops': depth}

    visited.add(key)
    sheet = wb[sheet_name]
    cell = sheet[cell_ref]

    if cell.data_type != 'f':
        return {'lines': [f"{'  ' * depth}{sheet_name}!{cell_ref} = {cell.value}"], 'hops': depth}

    trace_lines = [f"{'  ' * depth}{sheet_name}!{cell_ref} = {cell.value}"]
    pattern = r"(?:'([^']+)'!)?([A-Z]+[0-9]+)"
    matches = re.findall(pattern, cell.value)

    for ref_sheet, ref_cell in matches:
        actual_sheet = ref_sheet if ref_sheet else sheet_name
        if actual_sheet in wb.sheetnames:
            sub_trace = trace_formula_chain_with_hops(wb, actual_sheet, ref_cell, visited, depth + 1, max_depth)
            trace_lines.extend(sub_trace['lines'])

    return {'lines': trace_lines, 'hops': depth}
